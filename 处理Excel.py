#!python3
#-*- coding:utf-8 -*-
#更多请参考http://openpyxl.readthedocs.org
import openpyxl

#打开一个excel，只能打开当前目录的文档
wb=openpyxl.load_workbook('example.xlsx')

#调用get_sheet_names()方法取得工作表中所有表的名字
sheets=wb.get_sheet_names()
print(sheets)
print('1*'*20)

#get_sheet_by_name()
sheet=wb.get_sheet_by_name('Sheet1')
print(sheet['A1'].value)
c=sheet['B1']
print(c.value)
print('Row '+str(c.row)+',Column '+c.column+' is '+c.value)
print('Cell '+c.coordinate+' is ' +c.value)
print('2*'*20)

#cell()方法可以传入row和column参数，确定一个单元格
print(sheet.cell(row=1,column=2))
print(sheet.cell(row=1,column=2).value)
for i in range(1,8,2):
    print(i,sheet.cell(row=i,column=2).value)
print('3*'*20)

#max_row和max_column可以得到单元格的大小
print(sheet.max_row)
print(sheet.max_column)
print('4*'*20)

#列字母和数字之间的转换
from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(1))
print(column_index_from_string('AA'))
print('5*'*20)

#从表中取得行和列
for rowOfCellObjects in sheet['A1':'C3']:#也可以写成sheet.iter_cols(min_col=1,max_col=3,min_row=1,min_row=3)
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate,cellObj.value)
    print('--end of row--')
print('6*'*20)

#访问特定行或列的单元格的值
sheet=wb.get_active_sheet()
for cellObj in sheet['B']:
    print(cellObj.value)
